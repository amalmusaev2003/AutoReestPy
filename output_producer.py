import sys
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl
from parse_engine import *

class WorkerThread(QThread):
    progress_updated = pyqtSignal(int)
    log_updated = pyqtSignal(str)

    def __init__(self, path_in, path_out):
        super().__init__()
        self.path_in = path_in
        self.path_out = path_out

    def run(self):
        try:
            files = get_pdf_files(self.path_in)
            total_files = len(files)
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.column_dimensions['A'].width = 40
            worksheet.column_dimensions['B'].width = 30
            cell_a1 = worksheet['A1']
            cell_a1.value = "Номер последнего изменения документа"
            cell_b1 = worksheet['B1']
            cell_b1.value = "Обозначение документа"
            header = openpyxl.styles.Font(color="FF0000")
            hyperlink = openpyxl.styles.Font(underline='single', color='0563C1')
            cell_a1.font = header
            cell_b1.font = header
            row = 2
            for idx, file in enumerate(files, start=1):
                link = file
                try:
                    last_number = get_last_number(file)
                    doc_num = get_document_number(file)
                    if last_number == -1: #Смотреть в parse_engine.py
                        last_number = 0
                        self.log_updated.emit(f"{file} Файл не считался. AssertionError")
                    elif last_number == -2: #Смотреть в parse_engine.py
                        last_number = 0
                        self.log_updated.emit(f"{file} Номер последнего изменения считался с ошибкой")
                    elif last_number > 100:
                        last_number = 0
                except:
                    self.log_updated.emit(f"С файлами что-то не так, парсер не смог считать с них информацию...")
                    return
                
                cell_ai = worksheet.cell(row=row, column=1)
                cell_bi = worksheet.cell(row=row, column=2)

                cell_ai.value = f"{last_number}"
                cell_bi.value = '=HYPERLINK("{}", "{}")'.format(link, f"{doc_num}")
                cell_bi.font = hyperlink
                row += 1

                progress_percentage = idx * 100 // total_files
                self.progress_updated.emit(progress_percentage)
                self.log_updated.emit(f"Обработано файлов: {idx}/{total_files}: {last_number} - {doc_num} - {file}")
            workbook.save(self.path_out)
            workbook.close()
        except:
            self.log_updated.emit(f"Что-то пошло не так, попробуйте другой набор файлов")